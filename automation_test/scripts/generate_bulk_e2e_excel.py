#!/usr/bin/env python3
"""Generate Bulk Deployments E2E testcase matrix matching device_tags Excel layout."""
from __future__ import annotations

from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

TEMPLATE = Path.home() / "Downloads" / "device_tags_e2e_testcases_updated_actual_results.xlsx"
OUT = (
    Path(__file__).resolve().parent.parent
    / "tests"
    / "bulk-deployment-tests"
    / "bulk_deployments_e2e_testcases.xlsx"
)

HEADERS = (
    "Testcase ID",
    "Category",
    "Testcase Description",
    "Pre-condition",
    "Steps",
    "Expected results",
    "Actual results",
)

COL_WIDTHS = {"A": 18, "B": 18, "C": 36, "D": 34, "E": 52, "F": 48, "G": 62}

ROWS = [
    {
        "id": "TC-BULK-E2E-001",
        "category": "Core flow / Publish",
        "description": "Immediate publish — create draft, assign app and device, publish, verify successful finish and batches",
        "spec": "automation_test/tests/bulk-deployment-tests/bd-e2e-core-flow.test.js",
        "pre": (
            "1. User is authenticated.\n"
            "2. Bulk Deployments is reachable from the test environment configuration.\n"
            "3. Configured test apps and at least one online device search keyword (bulkDeployments) are available."
        ),
        "steps": (
            "1. Create draft and assign one configured app and one online device.\n"
            "2. Verify assigned resources on Apps and Devices tabs.\n"
            "3. Publish and wait until overview shows Completed or Published.\n"
            "4. On Devices tab, verify the device row shows Completed.\n"
            "5. Open Batches tab; verify batch metrics are non-negative numbers and failed count is 0."
        ),
        "expected": (
            "1. After publish completes, deployment status is **no longer Draft**; overview badge shows "
            "**Completed** or **Published** (successful publish).\n"
            "2. **Apps** tab: the assigned app appears in **exactly one row**; app metadata matches the selection.\n"
            "3. **Devices** tab: assigned device row is visible; per-device deployment status on that row is "
            "**Completed** (per current table layout).\n"
            "4. **Batches** tab: every metric is a **number** and **≥ 0**; for this run **failed batch count = 0**."
        ),
    },
    {
        "id": "TC-BULK-E2E-002",
        "category": "Core flow / Schedule",
        "description": "Future schedule — publish reaches Scheduled state with Start On set, list row matches",
        "spec": "automation_test/tests/bulk-deployment-tests/bd-e2e-core-flow.test.js",
        "pre": (
            "1. User is authenticated.\n"
            "2. Bulk Deployments environment and default future-schedule offset (days ahead) are valid for the test account."
        ),
        "steps": (
            "1. Create draft with future schedule metadata, one app, and one online device.\n"
            "2. Publish and wait until overview status is Scheduled.\n"
            "3. Read Start On on overview.\n"
            "4. Open list, search for the deployment name, verify row and Scheduled status."
        ),
        "expected": (
            "1. After publish, deployment status is **Scheduled** (no longer Draft).\n"
            "2. **Start On** on the detail overview is **populated** with a concrete date/time (non-empty, readable).\n"
            "3. Schedule-related fields on **detail** stay **consistent**; the **list** row for this deployment also "
            "reflects **Scheduled**."
        ),
    },
    {
        "id": "TC-BULK-E2E-003",
        "category": "Cleanup",
        "description": "Cleanup behavior — created draft can be deleted from detail",
        "spec": "automation_test/tests/bulk-deployment-tests/bd-e2e-core-flow.test.js",
        "pre": "1. User is authenticated.\n2. User may delete draft deployments.",
        "steps": (
            "1. Create draft deployment (detail open).\n"
            "2. Delete from detail with confirmation.\n"
            "3. Search list by name; if row still visible, delete from list; expect no results."
        ),
        "expected": (
            "1. After delete confirms, the deployment is **gone** as an active entity (user lands on list or equivalent).\n"
            "2. **Bulk Deployments list** shows **no row** for that deployment (including after refresh).\n"
            "3. Searching by the **deleted name** returns **no matching row** (empty / no-result state)."
        ),
    },
    {
        "id": "TC-BULK-E2E-004",
        "category": "Consistency",
        "description": "Device Deployments tab status matches Bulk Deployment Devices tab",
        "spec": "automation_test/tests/bulk-deployment-tests/bd-e2e-core-flow.test.js",
        "pre": (
            "1. User is authenticated.\n"
            "2. Flow can produce a failed bulk deployment with one assigned online device."
        ),
        "steps": (
            "1. Create and publish until bulk deployment status is Failed.\n"
            "2. Read device-level status from Bulk Deployment Devices tab.\n"
            "3. Open the device from Devices tab.\n"
            "4. On device detail, open Deployments tab and find the row for this deployment; status cell or row text matches."
        ),
        "expected": (
            "1. When the scenario completes, bulk deployment overview shows status **Failed**.\n"
            "2. On bulk deployment **Devices** tab, the assigned device row shows deployment status **Failed**.\n"
            "3. On **device detail → Deployments** tab, the row for this bulk deployment still shows **Failed** "
            "(same as bulk Devices tab)."
        ),
    },
    {
        "id": "TC-BULK-E2E-010",
        "category": "Duplicate",
        "description": "Duplicate twice from detail keeps valid Draft copies",
        "spec": "automation_test/tests/bulk-deployment-tests/bd-e2e-action-twice.test.js",
        "pre": "1. User is authenticated.\n2. Duplicate action is allowed on draft detail.",
        "steps": (
            "1. Create source draft with a fixed version (e.g. 3.3.3).\n"
            "2. Duplicate once; verify Draft and version preserved.\n"
            "3. Duplicate again; verify Draft and version preserved."
        ),
        "expected": (
            "1. Each **Duplicate** yields a **valid Draft** deployment (status badge **Draft**).\n"
            "2. On each duplicate copy, overview **Version** stays **3.3.3**.\n"
            "3. UI remains **stable** after two duplicates: no crash, dialogs close, overview remains readable."
        ),
    },
    {
        "id": "TC-BULK-E2E-011",
        "category": "Apps / Dedup",
        "description": "Assign same app twice keeps one row in Apps table",
        "spec": "automation_test/tests/bulk-deployment-tests/bd-e2e-action-twice.test.js",
        "pre": "1. User is authenticated.\n2. A known app name exists in Add App search.",
        "steps": (
            "1. Create draft with one app assigned.\n"
            "2. Open Add App and assign the same app again; confirm Assign.\n"
            "3. Open Apps tab; count tbody rows containing the app name."
        ),
        "expected": (
            "1. **Apps** table contains **only one row** for the selected app.\n"
            "2. A second assignment is **blocked** or **safely ignored**; no duplicate row is created.\n"
            "3. **No** second row for the same app; table data stays consistent."
        ),
    },
    {
        "id": "TC-BULK-E2E-012",
        "category": "Devices / Dedup",
        "description": "Assign same device twice keeps one row in Devices table",
        "spec": "automation_test/tests/bulk-deployment-tests/bd-e2e-action-twice.test.js",
        "pre": "1. User is authenticated.\n2. Online device search keyword resolves to a device.",
        "steps": (
            "1. Create draft with one device assigned.\n"
            "2. Reopen Add Device and add the same device again.\n"
            "3. Open Devices tab; count tbody rows containing the device keyword."
        ),
        "expected": (
            "1. **Devices** table contains **only one row** for the selected device.\n"
            "2. A second add is **blocked** or **safely ignored**; no duplicate row is created.\n"
            "3. **No** second row for the same device; table data stays consistent."
        ),
    },
    {
        "id": "TC-BULK-E2E-020",
        "category": "Version",
        "description": "Version older baseline persists across detail and list",
        "spec": "automation_test/tests/bulk-deployment-tests/bd-e2e-version-matrix.test.js",
        "pre": "1. User is authenticated.",
        "steps": (
            "1. Create draft with an older version string (e.g. 1.0.0) on overview.\n"
            "2. Open list, search deployment name, read version column cell."
        ),
        "expected": (
            "1. On **detail**, overview **Version** displays **1.0.0**.\n"
            "2. On **list**, the deployment row **Version** column shows **1.0.0**.\n"
            "3. After navigating list ↔ detail, **Version** stays **1.0.0** (no silent change)."
        ),
    },
    {
        "id": "TC-BULK-E2E-021",
        "category": "Version",
        "description": "Version equal (same version after edit save)",
        "spec": "automation_test/tests/bulk-deployment-tests/bd-e2e-version-matrix.test.js",
        "pre": "1. User is authenticated.",
        "steps": (
            "1. Create draft with baseline version.\n"
            "2. Open Edit, set version to the same value, save.\n"
            "3. Verify overview version unchanged."
        ),
        "expected": (
            "1. After edit-save with the same value, overview **Version** remains **2.2.2**.\n"
            "2. **No** unintended version bump or drop.\n"
            "3. Detail page stays **readable**; saved data is consistent with the entered version."
        ),
    },
    {
        "id": "TC-BULK-E2E-022",
        "category": "Version / Publish",
        "description": "Version newer survives publish and list row",
        "spec": "automation_test/tests/bulk-deployment-tests/bd-e2e-version-matrix.test.js",
        "pre": (
            "1. User is authenticated.\n"
            "2. App + online device fixtures support publish to a non-draft terminal or in-progress state."
        ),
        "steps": (
            "1. Create draft with assignments and a high/new version (e.g. 9.9.9).\n"
            "2. Publish.\n"
            "3. Verify detail overview still shows newer version.\n"
            "4. On list, search and verify version column includes newer value."
        ),
        "expected": (
            "1. After publish, deployment is **no longer Draft**; badge may be **Published**, **In Progress**, or "
            "**Completed** depending on timing.\n"
            "2. Overview **Version** stays **9.9.9**.\n"
            "3. On **list**, the deployment row **Version** column still shows **9.9.9**."
        ),
    },
    {
        "id": "TC-BULK-E2E-023",
        "category": "Version / Edit",
        "description": "Upgrade in-place on draft (older -> newer) via Edit",
        "spec": "automation_test/tests/bulk-deployment-tests/bd-e2e-version-matrix.test.js",
        "pre": "1. User is authenticated.",
        "steps": (
            "1. Create draft with older version.\n"
            "2. Open Edit, change version to newer, save.\n"
            "3. Verify overview shows newer version."
        ),
        "expected": (
            "1. Overview **Version** updates from **1.0.0** to **2.0.0**.\n"
            "2. **2.0.0** is shown correctly in **Overview** after save.\n"
            "3. Edit-save **completes** (modal dismissed, no blocking error); detail remains usable."
        ),
    },
    {
        "id": "TC-BULK-E2E-024",
        "category": "Version / Edit",
        "description": "Downgrade in-place on draft (newer -> older) via Edit",
        "spec": "automation_test/tests/bulk-deployment-tests/bd-e2e-version-matrix.test.js",
        "pre": "1. User is authenticated.",
        "steps": (
            "1. Create draft with newer version.\n"
            "2. Open Edit, change version to older, save.\n"
            "3. Verify overview shows older version."
        ),
        "expected": (
            "1. Overview **Version** updates from **5.0.0** to **3.0.0**.\n"
            "2. **3.0.0** is shown correctly in **Overview** after save.\n"
            "3. Edit-save **completes**; detail remains usable."
        ),
    },
    {
        "id": "TC-BULK-E2E-030",
        "category": "Apps",
        "description": "Add App valid -> table row visible",
        "spec": "automation_test/tests/bulk-deployment-tests/bd-e2e-app-device-matrix.test.js",
        "pre": "1. User is authenticated.\n2. Valid app name exists for Add App flow.",
        "steps": "1. Create draft.\n2. Add app by name via UI.\n3. Open Apps tab; locate row for app.",
        "expected": (
            "1. **Apps** tab shows **one visible row** for the selected app.\n"
            "2. App information in the row **matches** the catalog selection (name/build per UI).\n"
            "3. **Apps** tab layout remains **readable** after assignment (headers, actions intact)."
        ),
    },
    {
        "id": "TC-BULK-E2E-031",
        "category": "Apps / Search",
        "description": "Add App non-existent keyword -> empty state + Assign disabled",
        "spec": "automation_test/tests/bulk-deployment-tests/bd-e2e-app-device-matrix.test.js",
        "pre": "1. User is authenticated.",
        "steps": (
            "1. Create draft.\n"
            "2. Open Add App modal.\n"
            "3. Search a random non-matching keyword.\n"
            "4. Observe no-apps message and Assign button disabled."
        ),
        "expected": (
            "1. Add App modal shows a **no-match / empty-state** message for the invalid keyword.\n"
            "2. **Assign** is **disabled** (no app added from this search).\n"
            "3. After closing the modal, deployment **still has no extra app** from that failed search."
        ),
    },
    {
        "id": "TC-BULK-E2E-032",
        "category": "Devices",
        "description": "Add Device valid -> row visible and status shown",
        "spec": "automation_test/tests/bulk-deployment-tests/bd-e2e-app-device-matrix.test.js",
        "pre": "1. User is authenticated.\n2. Online device search keyword is configured.",
        "steps": "1. Create draft.\n2. Add device by name.\n3. Open Devices tab; verify device row (and status) per helper.",
        "expected": (
            "1. **Devices** tab shows a **clear row** for the selected device.\n"
            "2. **Deployment status** on that row has a **non-empty** value (per product UI).\n"
            "3. Device identifying fields **match** the chosen device."
        ),
    },
    {
        "id": "TC-BULK-E2E-033",
        "category": "Devices / Search",
        "description": "Add Device non-existent keyword -> empty state + Add disabled",
        "spec": "automation_test/tests/bulk-deployment-tests/bd-e2e-app-device-matrix.test.js",
        "pre": "1. User is authenticated.",
        "steps": (
            "1. Create draft.\n"
            "2. Open Add Device modal.\n"
            "3. Search random non-matching keyword.\n"
            "4. Observe no devices message and primary Add action disabled."
        ),
        "expected": (
            "1. Modal shows **no-devices-found / empty-state** messaging.\n"
            "2. Primary **Add** in the modal stays **disabled**.\n"
            "3. **No device** is added to the deployment from that search attempt."
        ),
    },
    {
        "id": "TC-BULK-E2E-040",
        "category": "Failure / Retry",
        "description": "Create a failed deployment from the UI flow, then retry from detail",
        "spec": "automation_test/tests/bulk-deployment-tests/bd-e2e-failure-retry.test.js",
        "pre": (
            "1. User is authenticated.\n"
            "2. Product flow can reach Failed status with configured app and online device."
        ),
        "steps": (
            "1. Create draft, publish, wait until status Failed.\n"
            "2. Navigate away to detail by id; confirm still Failed.\n"
            "3. Trigger Retry, confirm dialogs, wait for a follow-up status among In Progress, Failed, Completed, Published."
        ),
        "expected": (
            "1. After first publish, overview settles on **Failed** before Retry.\n"
            "2. After **Retry**, the UI is **not broken**: deployment name and primary actions remain readable.\n"
            "3. Within timeout, status badge moves to **In Progress**, **Failed**, **Completed**, or **Published** "
            "(no invalid stuck state)."
        ),
    },
    {
        "id": "TC-BULK-E2E-041",
        "category": "Failure / Actions",
        "description": "After Failed status, detail shows Retry, Edit, Delete and hides Run Deployment",
        "spec": "automation_test/tests/bulk-deployment-tests/bd-e2e-failure-retry.test.js",
        "pre": "1. User is authenticated.\n2. Flow can produce Failed deployment.",
        "steps": (
            "1. Create and publish until Failed.\n"
            "2. Assert Retry, Edit, Delete visible.\n"
            "3. Assert Run Deployment is hidden."
        ),
        "expected": (
            "1. While status is **Failed**, **Run Deployment** is **not shown** (hidden / not offered).\n"
            "2. **Retry**, **Edit**, and **Delete** are **visible** and remain **consistent** on the same Failed view.\n"
            "3. **Delete** and other controls are not obscured or duplicated in a misleading way."
        ),
    },
    {
        "id": "TC-BULK-E2E-042",
        "category": "Failure / Downgrade",
        "description": "Downgrade publish fails; Retry without editing plan still ends in Failed",
        "spec": "automation_test/tests/bulk-deployment-tests/bd-e2e-failure-retry.test.js",
        "pre": (
            "1. User is authenticated.\n"
            "2. e2eOlderCatalogAppName and e2eNewerCatalogAppName configured; device already has newer build installed.\n"
            "3. Test is skipped when catalog pair is not configured."
        ),
        "steps": (
            "1. Create and publish downgrade scenario until Failed.\n"
            "2. Retry without changing apps/devices.\n"
            "3. Wait until status Failed again."
        ),
        "expected": (
            "1. After downgrade publish, overview ends in **Failed**.\n"
            "2. **Retry** without plan changes returns overview to **Failed** again.\n"
            "3. If catalog fixtures are missing, automation **skips**; manual run: record **N/A / Blocked**."
        ),
    },
    {
        "id": "TC-BULK-E2E-043",
        "category": "Failure / Recovery",
        "description": "After downgrade failure, replace app with newer catalog build, Retry, reach successful finish",
        "spec": "automation_test/tests/bulk-deployment-tests/bd-e2e-failure-retry.test.js",
        "pre": (
            "1. User is authenticated.\n"
            "2. Same catalog downgrade preconditions as TC-BULK-E2E-042; test skipped if not configured."
        ),
        "steps": (
            "1. Fail deployment with older catalog app.\n"
            "2. On Apps tab remove older app and assign newer catalog app.\n"
            "3. Retry; wait for Completed or Published."
        ),
        "expected": (
            "1. After assigning the **newer** catalog app and **Retry**, overview reaches **Completed** or **Published**.\n"
            "2. Final state is **successful** (not stuck Failed on the old downgrade-only plan).\n"
            "3. If catalog preconditions are missing: **Skipped** / **N/A** as in TC-BULK-E2E-042."
        ),
    },
    {
        "id": "TC-BULK-E2E-050",
        "category": "List / Search",
        "description": "Search by deployment name and invalid keyword on list",
        "spec": "automation_test/tests/bulk-deployment-tests/bd-e2e-list-sort-search.test.js",
        "pre": "1. User is authenticated.",
        "steps": (
            "1. Create draft with unique name.\n"
            "2. Open list, search exact name, verify row.\n"
            "3. Search nonsense keyword; verify no-result state."
        ),
        "expected": (
            "1. Search with the **exact** deployment name: list shows that deployment row (or unambiguous match).\n"
            "2. Search with a **nonsense** keyword: list shows **empty / no-result** state.\n"
            "3. Search box value (and URL query if used) **reflects** the active keyword."
        ),
    },
    {
        "id": "TC-BULK-E2E-051",
        "category": "List / Sort",
        "description": "Sort toggles on all sortable list columns and excludes Actions",
        "spec": "automation_test/tests/bulk-deployment-tests/bd-e2e-list-sort-search.test.js",
        "pre": (
            "1. User is authenticated.\n"
            "2. List has at least one data row for sort URL assertions."
        ),
        "steps": (
            "1. Open list until tbody has rows.\n"
            "2. For each sortable column (Deployment name, Version, Start On, End On, Status): click header twice; URL sort/order reflect allowed fields.\n"
            "3. Click Actions column header; URL sort must not become actions; sort/order remain coherent."
        ),
        "expected": (
            "1. After each sortable-header interaction, URL has a **valid sort field** (name, version, date fields, "
            "status…) and **order asc or desc**.\n"
            "2. Clicking **Actions** never yields **sort=actions**; sort/order remains **coherent**.\n"
            "3. List still **renders table rows** when finished (no blank / error surface)."
        ),
    },
]

# Recorded from Playwright Chromium runs against automation_test dev config.
_RUN_META = "Recorded from Playwright runs on 2026-05-12 (Chromium, dev).\n\n"

_PASS = (
    _RUN_META
    + "Automated E2E executed with Playwright Chromium against the dev environment "
    "(config/environments/dev.js). Steps completed within configured timeouts.\n"
    "Result: Pass.\n"
    "Spec: {spec}"
)
_FAIL_001 = (
    _RUN_META
    + "Automated E2E ran with Playwright Chromium against dev. After publish, overview reached a "
    "successful terminal status, but the step \"Verify assigned device shows Completed\" failed: "
    "the device row matching search keyword 3576M showed deployment status Failed while the test "
    "expected Completed (row text included Failed Online).\n"
    "Result: Fail.\n"
    "Spec: {spec}"
)
_FAIL_011 = (
    _RUN_META
    + "Automated E2E ran with Playwright Chromium against dev. The test failed while reopening "
    "Add App to assign the same app again: locator.fill timed out waiting for "
    "getByPlaceholder('Search and select app') (30s).\n"
    "Result: Fail.\n"
    "Spec: {spec}"
)
_FAIL_012 = (
    _RUN_META
    + "Automated E2E ran with Playwright Chromium against dev. The test failed in Add Device when "
    "selecting the same device again: expect(locator).toBeVisible failed for "
    ".device-selector-selected-name filtered by 3576M (elements not found within 30s).\n"
    "Result: Fail.\n"
    "Spec: {spec}"
)
_SKIP_CATALOG = (
    _RUN_META
    + "Playwright did not execute this case: test.skip — e2eOlderCatalogAppName and "
    "e2eNewerCatalogAppName are not configured (or device/catalog preconditions missing) in "
    "bulk deployment E2E fixtures.\n"
    "Result: Skipped.\n"
    "Spec: {spec}"
)

ACTUAL_RESULTS = {
    "TC-BULK-E2E-001": _FAIL_001,
    "TC-BULK-E2E-002": _PASS,
    "TC-BULK-E2E-003": _PASS,
    "TC-BULK-E2E-004": _PASS,
    "TC-BULK-E2E-010": _PASS,
    "TC-BULK-E2E-011": _FAIL_011,
    "TC-BULK-E2E-012": _FAIL_012,
    "TC-BULK-E2E-020": _PASS,
    "TC-BULK-E2E-021": _PASS,
    "TC-BULK-E2E-022": _PASS,
    "TC-BULK-E2E-023": _PASS,
    "TC-BULK-E2E-024": _PASS,
    "TC-BULK-E2E-030": _PASS,
    "TC-BULK-E2E-031": _PASS,
    "TC-BULK-E2E-032": _PASS,
    "TC-BULK-E2E-033": _PASS,
    "TC-BULK-E2E-040": _PASS,
    "TC-BULK-E2E-041": _PASS,
    "TC-BULK-E2E-042": _SKIP_CATALOG,
    "TC-BULK-E2E-043": _SKIP_CATALOG,
    "TC-BULK-E2E-050": _PASS,
    "TC-BULK-E2E-051": _PASS,
}


def main() -> None:
    template_wb = openpyxl.load_workbook(TEMPLATE)
    src = template_wb["Device Tags E2E"]
    header_font = copy(src["A1"].font)
    header_align = copy(src["A1"].alignment)
    data_align = Alignment(wrap_text=True, vertical="top")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk Deployments E2E"

    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.alignment = header_align

    for i, row in enumerate(ROWS, start=2):
        template = ACTUAL_RESULTS.get(row["id"])
        if not template:
            template = (
                "Result not recorded in ACTUAL_RESULTS map; re-run generate_bulk_e2e_excel.py after updating.\n"
                "Spec: {spec}"
            )
        actual = template.format(spec=row["spec"])
        values = [
            row["id"],
            row["category"],
            row["description"],
            row["pre"],
            row["steps"],
            row["expected"],
            actual,
        ]
        body_font = Font(name="Carlito", size=11)
        for c, val in enumerate(values, start=1):
            cell = ws.cell(i, c, val)
            cell.alignment = data_align
            cell.font = body_font

    wb.save(OUT)
    template_wb.close()
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
