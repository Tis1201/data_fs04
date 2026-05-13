#!/usr/bin/env python3
"""Bulk Deployments manual testcase matrix (TC-BULK-* excluding TC-BULK-E2E-*) — same columns as device_tags template."""
from __future__ import annotations

import re
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

ROOT = Path(__file__).resolve().parent.parent
TESTS_DIR = ROOT / "tests" / "bulk-deployment-tests"
TEMPLATE = Path.home() / "Downloads" / "device_tags_e2e_testcases_updated_actual_results.xlsx"
OUT = TESTS_DIR / "bulk_deployments_manual_testcases.xlsx"

HEADERS = (
    "Testcase ID",
    "Category",
    "Testcase Description",
    "Pre-condition",
    "Steps",
    "Expected results",
    "Actual results",
)

COL_WIDTHS = {"A": 22, "B": 20, "C": 40, "D": 34, "E": 52, "F": 48, "G": 55}

CATEGORY_BY_AREA = {
    "INFO": "List / Detail / Tabs",
    "CREATE": "Create deployment",
    "VERSION": "Version",
    "APPS": "Apps",
    "DEVICES": "Devices",
    "BATCHES": "Batches",
    "DUPLICATE": "Duplicate",
    "EDIT": "Edit deployment",
    "DELETE": "Delete",
    "PUBLISH": "Publish",
}

AREA_ORDER = {
    "INFO": 0,
    "CREATE": 1,
    "VERSION": 2,
    "APPS": 3,
    "DEVICES": 4,
    "BATCHES": 5,
    "DUPLICATE": 6,
    "EDIT": 7,
    "DELETE": 8,
    "PUBLISH": 9,
}

TEST_CALL_RE = re.compile(r"test\(\s*['\"]([^'\"]+)['\"]")

FIRST_TC_RE = re.compile(r"(TC-BULK-[A-Z]+-\d+)")


def primary_sort_key(testcase_id: str) -> tuple[int, int, str]:
    m = FIRST_TC_RE.search(testcase_id)
    if not m:
        return (99, 9999, testcase_id)
    area_m = re.match(r"TC-BULK-([A-Z]+)-(\d+)", m.group(1))
    if not area_m:
        return (99, 9999, testcase_id)
    area, num = area_m.group(1), int(area_m.group(2))
    return (AREA_ORDER.get(area, 98), num, testcase_id)


def infer_category(testcase_id: str) -> str:
    m = FIRST_TC_RE.search(testcase_id)
    if not m:
        return "Bulk Deployments"
    area_m = re.match(r"TC-BULK-([A-Z]+)-", m.group(1))
    if not area_m:
        return "Bulk Deployments"
    return CATEGORY_BY_AREA.get(area_m.group(1), "Bulk Deployments")


def split_title(raw: str) -> tuple[str, str]:
    if ": " in raw:
        left, right = raw.split(": ", 1)
        return left.strip(), right.strip()
    return raw.strip(), ""


def collect_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for path in sorted(TESTS_DIR.glob("bd-*.test.js")):
        if path.name.startswith("bd-e2e-"):
            continue
        text = path.read_text(encoding="utf-8")
        rel_spec = f"automation_test/tests/bulk-deployment-tests/{path.name}"
        for m in TEST_CALL_RE.finditer(text):
            raw = m.group(1)
            if "TC-BULK-E2E-" in raw:
                continue
            testcase_id, description = split_title(raw)
            if not description:
                description = testcase_id
            pre = (
                "1. User is authenticated.\n"
                "2. User can access Bulk Deployments (list and deployment detail) in the target environment.\n"
                "3. Use valid test data (apps, devices, tags) as required by the scenario."
            )
            steps = (
                "1. Navigate to Bulk Deployments (list or open the relevant draft/detail) as appropriate.\n"
                f"2. Perform the actions described in this testcase: {description}.\n"
                "3. Observe UI feedback, field values, tables, modals, and status badges.\n"
                "4. Where applicable, confirm list and detail stay consistent after refresh or navigation."
            )
            expected = (
                f"1. The product behavior matches the intent of: {description}.\n"
                "2. Validation, empty states, and action availability follow current product rules "
                "(Draft vs Published/Scheduled/Failed, etc.).\n"
                "3. No unintended data loss except where the scenario explicitly deletes or cancels."
            )
            actual = (
                "Manual execution: to be recorded (tester, date, build/environment, Pass/Fail, notes).\n"
                "Automation reference (Playwright): "
                f"{rel_spec}\n"
                "Result: Pending."
            )
            rows.append(
                {
                    "id": testcase_id,
                    "category": infer_category(testcase_id),
                    "description": description,
                    "pre": pre,
                    "steps": steps,
                    "expected": expected,
                    "actual": actual,
                }
            )
    rows.sort(key=lambda r: primary_sort_key(r["id"]))
    return rows


def main() -> None:
    template_wb = openpyxl.load_workbook(TEMPLATE)
    src = template_wb["Device Tags E2E"]
    header_font = copy(src["A1"].font)
    header_align = copy(src["A1"].alignment)
    data_align = Alignment(wrap_text=True, vertical="top")
    body_font = Font(name="Carlito", size=11)

    rows = collect_rows()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk Deployments Manual"

    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.alignment = header_align

    for i, row in enumerate(rows, start=2):
        values = [
            row["id"],
            row["category"],
            row["description"],
            row["pre"],
            row["steps"],
            row["expected"],
            row["actual"],
        ]
        for c, val in enumerate(values, start=1):
            cell = ws.cell(i, c, val)
            cell.alignment = data_align
            cell.font = body_font

    wb.save(OUT)
    template_wb.close()
    print(f"Wrote {OUT} ({len(rows)} rows)")


if __name__ == "__main__":
    main()
